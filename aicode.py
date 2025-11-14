# aicode.py - پردازش SEO هوشمند (آفلاین + Gemini اختیاری)
import pandas as pd
import re
import json
import time
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def run_seo(filepath, GEMINI_API_KEY=""):
    """
    filepath: مسیر فایل اکسل آپلود شده
    GEMINI_API_KEY: کلید API Gemini (اختیاری)
    خروجی: مسیر فایل اکسل تولید شده
    """
    # ------------------------------
    # 1. خواندن فایل اکسل
    # ------------------------------
    df = pd.read_excel(filepath, usecols=[0,1])
    df.columns = ['عبارت', 'حجم جستجو']
    df = df.dropna()
    df['حجم جستجو'] = pd.to_numeric(df['حجم جستجو'], errors='coerce')
    df = df.dropna()

    # ------------------------------
    # 2. تمیز کردن عبارات
    # ------------------------------
    def تمیز(متن):
        متن = str(متن)
        متن = re.sub(r'[\u200c\u200d\u200e\u200f]', '', متن)
        متن = متن.replace('-', ' ').replace('_', ' ').replace('/', ' ')
        متن = re.sub(r'\s+', ' ', متن).strip().lower()
        return متن if متن else "نامشخص"

    df['تمیز'] = df['عبارت'].apply(تمیز)
    df = df[df['تمیز'] != "نامشخص"].copy()

    # ------------------------------
    # 3. ادغام هوشمند عبارات مشابه
    # ------------------------------
    def مشابه(س1, س2):
        return SequenceMatcher(None, س1, س2).ratio() > 0.85

    parent = list(range(len(df)))
    rank = [0]*len(df)

    def find(x):
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]

    def union(x,y):
        px, py = find(x), find(y)
        if px != py:
            if rank[px] < rank[py]: parent[px] = py
            elif rank[px] > rank[py]: parent[py] = px
            else:
                parent[py] = px
                rank[px] += 1

    for i in range(len(df)):
        for j in range(i+1,len(df)):
            if مشابه(df['تمیز'].iloc[i], df['تمیز'].iloc[j]):
                union(i,j)

    groups = defaultdict(list)
    for i in range(len(df)):
        groups[find(i)].append(i)

    ادغام_شده = 0
    new_rows = []
    for root, group in groups.items():
        if len(group) >1: ادغام_شده += len(group)-1
        total_volume = sum(df['حجم جستجو'].iloc[k] for k in group)
        main_idx = max(group, key=lambda k: df['حجم جستجو'].iloc[k])
        main_original = df['عبارت'].iloc[main_idx]
        main_clean = df['تمیز'].iloc[main_idx]
        new_rows.append({
            'عبارت': main_original,
            'حجم جستجو': total_volume,
            'تمیز': main_clean
        })

    df = pd.DataFrame(new_rows)

    # ------------------------------
    # 4. Gemini یا حالت آفلاین
    # ------------------------------
    استفاده_از_gemini = False
    نتایج = []

    if GEMINI_API_KEY.startswith("AIzaSy"):
        try:
            import google.generativeai as genai
            genai.configure(api_key=GEMINI_API_KEY)
            model = genai.GenerativeModel('gemini-2.0-flash')
            model.generate_content("تست")
            استفاده_از_gemini = True
        except:
            استفاده_از_gemini = False

    if استفاده_از_gemini:
        BATCH_SIZE = 5
        for i in range(0,len(df),BATCH_SIZE):
            دسته = df['تمیز'].iloc[i:i+BATCH_SIZE].tolist()
            پرامپت = f"""
تحلیل کن و فقط JSON بده:
{chr(10).join([f'{j+1}. "{ع}"' for j, ع in enumerate(دسته)])}
فرمت:
[
  {{"intent": "Informational|Commercial|Transactional|Navigational", "title": "عنوان H1 (حداکثر 60 کاراکتر)"}}
]
"""
            try:
                پاسخ = model.generate_content(پرامپت)
                نتیجه = پاسخ.text.strip()
                if "```json" in نتیجه:
                    نتیجه = نتیجه.split("```json")[1].split("```")[0].strip()
                elif "```" in نتیجه:
                    نتیجه = نتیجه.split("```")[1].strip()
                تحلیل = json.loads(نتیجه)
                if len(تحلیل)!=len(دسته):
                    تحلیل = [{"intent":"Commercial","title":ع} for ع in دسته]
                نتایج.extend(تحلیل)
            except:
                نتایج.extend([{"intent":"Commercial","title":ع} for ع in دسته])
            time.sleep(6)
        df['Intent_AI'] = [x['intent'] for x in نتایج]
        df['H1_پیشنهادی'] = [x['title'] for x in نتایج]
    else:
        # حالت آفلاین
        کلمات = {
            "Transactional": ['خرید','قیمت','فروش','ارزان','تخفیف','اقساط','کد تخفیف'],
            "Commercial": ['بهترین','برترین','مقایسه','نقد','بررسی','رتبه','لیست'],
            "Informational": ['طرز تهیه','چگونه','آموزش','روش','فیلم','چیست','معنی'],
            "Navigational": ['سایت','ورود','اپلیکیشن','دانلود','دیجی کالا','آپارات']
        }

        def intent_ai(متن):
            متن = متن.lower()
            for intent, لیست in کلمات.items():
                if any(ک in متن for ک in لیست): return intent
            n = len(متن.split())
            if n>=5: return "Informational"
            elif n<=2: return "Navigational"
            else: return "Commercial"

        def h1_ai(متن):
            if 'طرز' in متن: return f"طرز تهیه {متن.replace('طرز تهیه','').strip()} در خانه"
            if 'بهترین' in متن: return f"{متن} + مقایسه 1404"
            if 'خرید' in متن: return f"خرید {متن.replace('خرید','').strip()} با گارانتی"
            if 'قیمت' in متن: return f"قیمت {متن.replace('قیمت','').strip()} امروز"
            return متن.title()[:60]

        df['Intent_AI'] = df['تمیز'].apply(intent_ai)
        df['H1_پیشنهادی'] = df['تمیز'].apply(h1_ai)

    # ------------------------------
    # 5. دسته‌بندی + Page Type + خروجی
    # ------------------------------
    df['برای_دسته'] = df['تمیز'].str.replace(r'^(طرز تهیه|خرید|قیمت|بهترین|چگونه)\s*','',regex=True,case=False)
    سه_کلمه = df['برای_دسته'].str.split().apply(lambda x:' '.join(x[:3]) if len(x)>=3 else None).dropna()
    تکرار_سه = [k for k,v in Counter(سه_کلمه).items() if v>=2]
    دو_کلمه = df['برای_دسته'].str.split().apply(lambda x:' '.join(x[:2]) if len(x)>=2 else None).dropna()
    تکرار_دو = [k for k,v in Counter(دو_کلمه).items() if v>=3]
    کلید_قوی = تکرار_سه + [d for d in تکرار_دو if not any(d in t for t in تکرار_سه)]

    def تخصیص_دسته(متن):
        if not متن: return "نامشخص"
        متن = متن.lower().strip()
        for کلید in کلید_قوی:
            if متن.startswith(کلید):
                return ' '.join(کلید.split()[:2]).title()
        کلمات = متن.split()
        return کلمات[0].title() if کلمات else "نامشخص"

    df['دسته'] = df['برای_دسته'].apply(تخصیص_دسته)
    df = df[df['دسته']!="نامشخص"].copy()

    خلاصه = df.groupby(['دسته','عبارت','Intent_AI'])['حجم جستجو'].sum().reset_index()
    جمع_دسته = خلاصه.groupby('دسته')['حجم جستجو'].sum().to_dict()

    def page_type(حجم,دسته):
        if حجم==جمع_دسته.get(دسته,0):
            if حجم>=100000: return "Pillar"
            elif حجم>=50000: return "Cluster"
        return "Sub-Cluster"

    خلاصه['Page Type'] = خلاصه.apply(lambda x: page_type(x['حجم جستجو'],x['دسته']),axis=1)

    داده = []
    رنگ = []

    for (دسته,intent), گروه in خلاصه.groupby(['دسته','Intent_AI']):
        جمع = گروه['حجم جستجو'].sum()
        pt = page_type(جمع,دسته)
        داده.append([دسته,جمع,intent,pt,""])
        رنگ.append(len(داده)-1)
        for _,r in گروه.iterrows():
            h1 = df[df['عبارت']==r['عبارت']]['H1_پیشنهادی'].iloc[0]
            داده.append([r['عبارت'],r['حجم جستجو'],intent,r['Page Type'],h1])

    جدول = pd.DataFrame(داده,columns=['دسته / عبارت','حجم جستجو','Intent_AI','Page Type','H1 پیشنهادی'])

    # ------------------------------
    # ذخیره فایل خروجی
    # ------------------------------
    import os
    os.makedirs("SEO_Output", exist_ok=True)
    output_file = f"SEO_Output/نتایج_هوشمند.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "SEO AI"

    رنگ_intent = {'Informational':'FFD965','Navigational':'A9D08E','Commercial':'8FAADC','Transactional':'F4B084'}

    for i,h in enumerate(جدول.columns,1):
        c = ws.cell(1,i,h)
        c.font = Font(bold=True,color="FFFFFF")
        c.fill = PatternFill("solid","4472C4")

    for i,row in enumerate(dataframe_to_rows(جدول,index=False,header=False),2):
        for j,val in enumerate(row,1):
            ws.cell(i,j,val)
        if (i-2) in رنگ:
            fill = PatternFill("solid", رنگ_intent.get(row[2],"D9D9D9"))
            for j in [1,2]:
                cell = ws.cell(i,j)
                cell.fill = fill
                cell.font = Font(bold=True,size=13)

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 50
    ws.freeze_panes = 'A2'

    wb.save(output_file)
    return output_file, ادغام_شده, len(خلاصه['دسته'].unique()), "Gemini" if استفاده_از_gemini else "Offline"
